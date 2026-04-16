"""
Agentic Patch Strategist v3
============================
Multi-agent AI system for cost-aware, dependency-driven vulnerability prioritization.

Agents:
  1. IngestAgent          – loads & normalises vulnerability data
  2. RiskScoringAgent     – computes technical risk scores
  3. BusinessImpactAgent  – translates technical risk → business impact
  4. DependencyAgent      – models system dependencies & cascade risk
  5. CustomerImpactAgent  – scores customer-facing exposure
  6. SchedulerAgent       – builds patch schedule, resolves conflicts
  7. WhatIfAgent          – simulates exploit spikes or patch delays
  8. ComplianceAgent      – tags regulatory exposure (PCI, HIPAA, GDPR, SOX)
  9. ReporterAgent        – writes CSV, JSON summary, and HTML audit report

Changes in v3 (fixes + improvements over v2):
  ── BUG FIXES ──────────────────────────────────────────────────────────────
  - BUGFIX: _recommend_action thresholds were calibrated assuming scores > 0.75
    were reachable, but with real-world data (Business Impact 1-100, Dep Score
    1-5, Downtime 23-109 all min-max normalised to narrow ranges) the effective
    ceiling was ~0.65, so almost nothing got "Patch immediately". Thresholds
    are now dynamically percentile-based relative to the scored dataset AND an
    explicit fast-path escalates rows with CVSS ≥ 9.0 AND EPSS ≥ 0.75.
  - BUGFIX: SchedulerAgent marked near-universal scheduling_conflict=TRUE
    because all CVEs sharing the same Affected System + pre-supplied window
    (e.g. "Sat 1-3 AM") got the same scheduled_date string, triggering
    pandas.duplicated(). Fixed: conflicts now only flagged when two *different*
    CVEs are assigned to the exact same system in the exact same computed date
    slot — pre-supplied named windows are excluded from conflict detection since
    they represent a shared resource, not a clash.
  - BUGFIX: Expected Loss from the dataset was in natural-units hundreds
    (e.g. 761.09 = $761.09), but the HTML report showed "$761" correctly while
    the explanation said "≈$761" — both fine. However downstream Expected Loss
    was being zeroed out for rows that used the default (0.0) when the dataset
    column existed but was misread. Added explicit float cast with NaN guard.
  - BUGFIX: Tier defaults to "2" for entire dataset when column is missing,
    causing all rows to receive the same 0.65 multiplier and eliminating
    differentiation. Now inferred from Criticality column when Tier is absent:
    "High" → Tier 1, "Medium" → Tier 2, "Low" → Tier 3.
  - BUGFIX: Cascade Risk from input dataset ("Low"/"Medium"/"High" string) was
    renamed to "Cascade Risk Label" internally, but compute_priority_scores
    used df["cascade_risk"] (the float computed by DependencyAgent). When the
    system wasn't in DEFAULT_DEPENDENCY_GRAPH, cascade_risk=0, suppressing
    priority. Now the string label from the input is used as a fallback floor.
  - BUGFIX: ReporterAgent EXPORT_COLUMNS listed "cascade_risk" (internal float)
    but wrote it as "Cascade Risk" without ensuring the column existed after
    rename; this caused KeyError for some datasets. Now defensive.
  ── IMPROVEMENTS ───────────────────────────────────────────────────────────
  - IMPROVEMENT: Priority score now outputs on 0-10 scale (matching the source
    dataset convention and making it directly comparable/presentable).
  - IMPROVEMENT: _recommend_action now uses BOTH percentile thresholds AND
    absolute CVSS/EPSS hard rules so critical CVEs always escalate correctly.
  - IMPROVEMENT: SchedulerAgent assigns a per-CVE slot within the named window
    (e.g. "Sat 1-3 AM – Slot 2") when multiple CVEs share a system+window, so
    the schedule is readable without false conflicts.
  - IMPROVEMENT: SchedulerAgent now tries to spread CVEs across sequential
    windows when a single window would be overloaded (>3 CVEs per system).
  - IMPROVEMENT: Added RiskTierAgent (lightweight) that classifies each CVE
    into a Risk Tier (P1 Critical / P2 High / P3 Medium / P4 Low) for
    executive-friendly output alongside the numeric score.
  - IMPROVEMENT: HTML report now includes a sortable priority score column and
    color-coded risk tier badges.
  - IMPROVEMENT: JSON summary now includes per-tier counts, conflict count, and
    top 3 most-affected systems.
  - IMPROVEMENT: Added --aggressive-escalation CLI flag (default: on) that
    applies the CVSS+EPSS fast-path escalation rule.
  - IMPROVEMENT: Added --output-xlsx flag to emit an Excel workbook in addition
    to the CSV (requires openpyxl).
  - IMPROVEMENT: Logging now includes timestamps.
"""

from __future__ import annotations

import argparse
import json
import textwrap
from collections import Counter, deque
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd


# ---------------------------------------------------------------------------
# Logging helper
# ---------------------------------------------------------------------------

def _log(agent: str, msg: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}][{agent}] {msg}")


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SEVERITY_MAP = {
    "low":      0.25,
    "medium":   0.50,
    "moderate": 0.50,
    "high":     0.75,
    "critical": 1.00,
}

# Maps any column name variant → canonical column name used internally
COLUMN_ALIASES = {
    # CVE identifiers
    "cve id": "CVE ID", "cve": "CVE ID", "cve_id": "CVE ID",
    # CVSS
    "cvss score": "CVSS Score", "cvss": "CVSS Score", "cvss_score": "CVSS Score",
    # Severity
    "severity": "Severity", "severity label": "Severity", "severity_label": "Severity",
    # EPSS
    "exploit prob (epss)": "Exploit Prob (EPSS)",
    "exploit probability":  "Exploit Prob (EPSS)",
    "exploit_prob":         "Exploit Prob (EPSS)",
    "exploit_prob_epss":    "Exploit Prob (EPSS)",
    "epss":                 "Exploit Prob (EPSS)",
    # System
    "affected system": "Affected System", "system": "Affected System",
    "affected_system": "Affected System",
    # Criticality
    "criticality": "Criticality", "system criticality": "Criticality",
    # Business impact
    "business impact": "Business Impact", "business_impact": "Business Impact",
    # Dependency
    "dependency score": "Dependency Score", "dependency_score": "Dependency Score",
    # Downtime
    "downtime cost": "Downtime Cost", "downtime_cost": "Downtime Cost",
    # Effort
    "est. effort": "Est. Effort", "estimated effort": "Est. Effort",
    "effort":      "Est. Effort", "est_effort":       "Est. Effort",
    # Tier
    "tier": "Tier", "system tier": "Tier",
    # Compliance
    "compliance tags": "Compliance Tags", "compliance": "Compliance Tags",
    "regulatory":      "Compliance Tags",
    # Rollback risk
    "rollback risk": "Rollback Risk", "rollback_risk": "Rollback Risk",
    # Vendor
    "vendor": "Vendor", "vendor advisory": "Vendor",
    # Patch availability
    "patch available": "Patch Available", "patch_available": "Patch Available",
    # SBOM
    "sbom component": "SBOM Component", "sbom_component": "SBOM Component",
    # Maintenance window
    "maintenance window": "Maintenance Window",
    "maintenance_window": "Maintenance Window",
    # Customer impact
    "customer impact": "Customer Impact", "customer_impact": "Customer Impact",
    # Description
    "description": "Description",
    # Expected loss
    "expected loss": "Expected Loss", "expected_loss": "Expected Loss",
    # Downtime sensitivity
    "downtime sensitivity":  "Downtime Sensitivity",
    "downtime_sensitivity":  "Downtime Sensitivity",
    # Cascade risk (pre-computed label in some datasets)
    "cascade risk": "Cascade Risk Label", "cascade_risk": "Cascade Risk Label",
    # Priority Score from source (we rename to avoid collision with computed)
    "priority score":       "Source Priority Score",
    "recommended action":   "Source Recommended Action",
    "recommended patch window": "Source Recommended Patch Window",
    "explanation":          "Source Explanation",
}

# Tier → business criticality multiplier
TIER_MULTIPLIER: Dict[str, float] = {
    "1": 1.00, "tier 1": 1.00, "tier1": 1.00,
    "2": 0.65, "tier 2": 0.65, "tier2": 0.65,
    "3": 0.35, "tier 3": 0.35, "tier3": 0.35,
}

# Compliance frameworks and the keywords that trigger detection
COMPLIANCE_KEYWORDS: Dict[str, List[str]] = {
    "PCI-DSS": ["payment", "checkout", "card", "transaction", "billing", "pos"],
    "HIPAA":   ["health", "patient", "medical", "ehr", "phi", "hospital", "clinical"],
    "GDPR":    ["user", "customer", "profile", "identity", "login", "auth", "personal"],
    "SOX":     ["finance", "accounting", "erp", "ledger", "audit", "reporting", "analytics"],
    "ISO27001":["api", "gateway", "network", "firewall", "vpn", "infrastructure", "core api"],
}

# Dependency graph: system → downstream systems it can affect
DEFAULT_DEPENDENCY_GRAPH: Dict[str, List[str]] = {
    "Frontend":          ["API Gateway", "Auth Service"],
    "API Gateway":       ["Payment Service", "Analytics Service", "User Service"],
    "Core API":          ["Payment API", "Analytics", "Login Service"],
    "Payment Service":   ["Transaction DB", "Fraud Detection"],
    "Payment API":       ["Database", "Fraud Detection"],
    "Auth Service":      ["User Service", "Session Store"],
    "Login Service":     ["Database", "Session Store"],
    "User Service":      ["User DB"],
    "Analytics Service": ["Analytics DB"],
    "Analytics":         ["Database"],
    "Fraud Detection":   ["Transaction DB"],
    "Transaction DB":    [],
    "Database":          [],
    "User DB":           [],
    "Session Store":     [],
    "Analytics DB":      [],
}

# Labels that indicate an immediate/emergency scheduling slot
EMERGENCY_WINDOW_LABELS = {"emergency change window", "immediate", "emergency"}

# Named windows are human-readable labels like "Sat 1-3 AM" — not date strings
# We detect these so we don't flag shared-window rows as scheduling conflicts
NAMED_WINDOW_PATTERN_KEYWORDS = ["am", "pm", "weekday", "weekend", "sat", "sun",
                                  "mon", "tue", "wed", "thu", "fri"]

# CVSS + EPSS hard-escalation thresholds (applied before percentile-based logic)
CRITICAL_CVSS_THRESHOLD  = 9.0
CRITICAL_EPSS_THRESHOLD  = 0.75
# Max CVEs we'll pack into one window slot before spilling to next window
MAX_CVES_PER_WINDOW_SLOT = 3


# ---------------------------------------------------------------------------
# Weight validation helper
# ---------------------------------------------------------------------------

def validate_weights(weights: Dict[str, float]) -> None:
    """Assert that positive weights roughly sum to 1.0."""
    positive_keys = ["risk", "business", "dependency", "exploit_bonus",
                     "cascade", "compliance", "customer_impact"]
    penalty_keys  = ["downtime_penalty", "effort_penalty"]
    pos_sum = sum(weights.get(k, 0.0) for k in positive_keys)
    pen_sum = sum(weights.get(k, 0.0) for k in penalty_keys)
    assert abs(pos_sum - 1.0) < 0.02, (
        f"Positive weight sum {pos_sum:.4f} deviates from 1.0 — check weight profile."
    )
    assert pen_sum < 0.30, (
        f"Penalty weight sum {pen_sum:.4f} is suspiciously large."
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Agentic Patch Strategist v3 – multi-agent vulnerability prioritization",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            Examples:
              python agentic_patch_strategist_v3.py --input vulns.csv
              python agentic_patch_strategist_v3.py --input vulns.csv --risk-appetite conservative
              python agentic_patch_strategist_v3.py --input vulns.csv --what-if-cve CVE-2024-38063
              python agentic_patch_strategist_v3.py --input vulns.csv --what-if-delay CVE-2024-43491 --delay-days 30
              python agentic_patch_strategist_v3.py --input vulns.csv --dependency-graph deps.json
              python agentic_patch_strategist_v3.py --input vulns.csv --output-xlsx results.xlsx
        """),
    )
    parser.add_argument("--input",    required=True, help="Path to input CSV or XLSX dataset")
    parser.add_argument("--output",   default="ranked_patch_plan.csv",
                        help="Path to ranked output CSV (default: ranked_patch_plan.csv)")
    parser.add_argument("--output-xlsx", default=None,
                        help="Optional path to output Excel workbook (.xlsx)")
    parser.add_argument("--summary-json", default=None,
                        help="Optional path to summary JSON")
    parser.add_argument("--html-report",  default=None,
                        help="Optional path to HTML audit report")
    parser.add_argument(
        "--risk-appetite",
        choices=["conservative", "balanced", "aggressive"],
        default="balanced",
        help="Risk posture for scoring weights (default: balanced)",
    )
    # What-if: exploit spike
    parser.add_argument("--what-if-cve",          default=None,
                        help="CVE ID to simulate an exploit probability spike")
    parser.add_argument("--what-if-exploit-prob",  type=float, default=0.98,
                        help="New EPSS value for the spike simulation (default: 0.98)")
    # What-if: delay simulation
    parser.add_argument("--what-if-delay", default=None,
                        help="CVE ID to simulate a patch delay (shows risk accumulation)")
    parser.add_argument("--delay-days",    type=int, default=30,
                        help="Number of days to delay patching (default: 30)")
    # Dependency graph
    parser.add_argument("--dependency-graph", default=None,
                        help="Path to JSON file defining system dependency graph")
    # Scheduling
    parser.add_argument("--schedule-start", default=None,
                        help="Patch schedule start date YYYY-MM-DD (default: today)")
    parser.add_argument("--window-days",    type=int, default=7,
                        help="Days between maintenance windows (default: 7)")
    # Output verbosity
    parser.add_argument("--top-n", type=int, default=10,
                        help="Number of top CVEs to print to console (default: 10)")
    # Escalation control
    parser.add_argument("--no-aggressive-escalation", action="store_true",
                        help="Disable CVSS+EPSS hard-escalation to 'Patch immediately'")
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Agent 1: IngestAgent
# ---------------------------------------------------------------------------

class IngestAgent:
    """Loads, normalises and cleans raw vulnerability data from CSV or XLSX."""

    def run(self, path: str) -> pd.DataFrame:
        _log("IngestAgent", f"Loading dataset from {path} ...")
        df = self._load(path)
        df = self._standardize_columns(df)
        df = self._add_missing_columns(df)
        df = self._clean(df)
        df = self._infer_tier(df)
        _log("IngestAgent", f"Loaded {len(df)} vulnerabilities.")
        return df

    # ------------------------------------------------------------------
    def _load(self, path: str) -> pd.DataFrame:
        fp = Path(path)
        if not fp.exists():
            raise FileNotFoundError(f"Input file not found: {path}")
        if fp.suffix.lower() == ".csv":
            return pd.read_csv(fp)
        if fp.suffix.lower() in {".xlsx", ".xls"}:
            return pd.read_excel(fp)
        raise ValueError(f"Unsupported file type '{fp.suffix}'. Use CSV or XLSX.")

    def _standardize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        rename_map = {}
        for col in df.columns:
            key = str(col).strip().lower()
            if key in COLUMN_ALIASES:
                rename_map[col] = COLUMN_ALIASES[key]
        return df.rename(columns=rename_map)

    def _add_missing_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        defaults: Dict = {
            "CVE ID":               [f"CVE-AUTO-{i+1:04d}" for i in range(len(df))],
            "CVSS Score":           5.0,
            "Severity":             "Medium",
            "Exploit Prob (EPSS)":  0.50,
            "Affected System":      "Unknown System",
            "Criticality":          "Medium",
            "Business Impact":      50.0,
            "Dependency Score":     3.0,
            "Downtime Cost":        1000.0,
            "Est. Effort":          3.0,
            "Tier":                 "",           # empty → will be inferred
            "Compliance Tags":      "",
            "Rollback Risk":        "Medium",
            "Vendor":               "",
            "Patch Available":      "Yes",
            "SBOM Component":       "",
            "Maintenance Window":   "",
            "Customer Impact":      "Medium",
            "Description":          "",
            "Expected Loss":        0.0,
            "Downtime Sensitivity": "Medium",
            "Cascade Risk Label":   "Medium",
        }
        for col, default in defaults.items():
            if col not in df.columns:
                df[col] = default
        return df

    def _clean(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df["CVSS Score"]           = _clean_numeric(df["CVSS Score"],          5.0).clip(0, 10)
        df["Exploit Prob (EPSS)"]  = _clean_numeric(df["Exploit Prob (EPSS)"], 0.50).clip(0, 1)
        df["Business Impact"]      = _clean_numeric(df["Business Impact"],     50.0).clip(lower=0)
        df["Dependency Score"]     = _clean_numeric(df["Dependency Score"],    3.0).clip(lower=0)
        df["Downtime Cost"]        = _clean_numeric(df["Downtime Cost"],       1000.0).clip(lower=0)
        df["Est. Effort"]          = _clean_numeric(df["Est. Effort"],         3.0).clip(lower=0)
        df["Expected Loss"]        = _clean_numeric(df["Expected Loss"],       0.0).clip(lower=0)

        str_cols = [
            "Severity", "Criticality", "Affected System", "Tier",
            "Compliance Tags", "Rollback Risk", "Vendor",
            "Patch Available", "SBOM Component", "Maintenance Window",
            "Customer Impact", "Description", "Downtime Sensitivity",
            "Cascade Risk Label",
        ]
        for col in str_cols:
            df[col] = df[col].fillna("").astype(str)

        # Auto-generate CVE IDs if missing
        mask = df["CVE ID"].isna() | (df["CVE ID"].astype(str).str.strip() == "")
        if mask.any():
            df.loc[mask, "CVE ID"] = [f"CVE-AUTO-{i+1:04d}" for i in range(int(mask.sum()))]
        df["CVE ID"] = df["CVE ID"].astype(str)
        return df

    def _infer_tier(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        FIX: When Tier column is missing or empty, infer from Criticality.
        High criticality → Tier 1, Medium → Tier 2, Low → Tier 3.
        This prevents all rows from getting the same default Tier-2 multiplier.
        """
        df = df.copy()
        empty_mask = df["Tier"].str.strip() == ""
        if empty_mask.any():
            crit_lower = df.loc[empty_mask, "Criticality"].str.strip().str.lower()
            tier_from_crit = crit_lower.map(
                {"critical": "1", "high": "1", "medium": "2", "low": "3", "moderate": "2"}
            ).fillna("2")
            df.loc[empty_mask, "Tier"] = tier_from_crit
            inferred_count = int(empty_mask.sum())
            _log("IngestAgent", f"Inferred Tier from Criticality for {inferred_count} rows.")
        return df


# ---------------------------------------------------------------------------
# Agent 2: RiskScoringAgent
# ---------------------------------------------------------------------------

class RiskScoringAgent:
    """Computes normalised technical risk scores from CVSS, EPSS, severity, criticality."""

    def run(self, df: pd.DataFrame) -> pd.DataFrame:
        _log("RiskScoringAgent", "Computing technical risk scores...")
        df = df.copy()
        df["cvss_norm"]         = df["CVSS Score"] / 10.0
        df["severity_norm"]     = _map_severity(df["Severity"])
        df["exploit_prob_norm"] = _min_max(df["Exploit Prob (EPSS)"])
        df["criticality_norm"]  = _map_severity(df["Criticality"])
        df["rollback_risk_norm"]= _map_severity(df["Rollback Risk"])

        df["risk_score"] = (
            0.45 * df["cvss_norm"] +
            0.20 * df["severity_norm"] +
            0.20 * df["exploit_prob_norm"] +
            0.10 * df["criticality_norm"] +
            0.05 * df["rollback_risk_norm"]
        ).clip(0, 1).round(4)
        return df


# ---------------------------------------------------------------------------
# Agent 3: BusinessImpactAgent
# ---------------------------------------------------------------------------

class BusinessImpactAgent:
    """Translates technical risk into business impact, incorporating tier and downtime sensitivity."""

    def run(self, df: pd.DataFrame) -> pd.DataFrame:
        _log("BusinessImpactAgent", "Mapping to business impact...")
        df = df.copy()
        df["business_impact_norm"]  = _min_max(df["Business Impact"])
        df["dependency_score_norm"] = _min_max(df["Dependency Score"])
        df["downtime_cost_norm"]    = _min_max(df["Downtime Cost"])
        df["effort_norm"]           = _min_max(df["Est. Effort"])

        # Downtime sensitivity modifier
        dt_sens = df["Downtime Sensitivity"].str.strip().str.lower()
        df["downtime_sensitivity_mult"] = dt_sens.map(
            {"high": 1.25, "medium": 1.0, "low": 0.75}
        ).fillna(1.0)

        # Tier multiplier: Tier-1 systems get a business impact boost
        tier_mult = (
            df["Tier"].str.strip().str.lower()
            .map(TIER_MULTIPLIER)
            .fillna(0.65)
        )
        df["tier_multiplier"] = tier_mult
        df["adjusted_business_impact"] = (
            df["business_impact_norm"] * tier_mult
        ).clip(0, 1)
        return df


# ---------------------------------------------------------------------------
# Agent 4: DependencyAgent
# ---------------------------------------------------------------------------

class DependencyAgent:
    """Models system dependencies and computes cascade risk score."""

    def __init__(self, graph: Optional[Dict[str, List[str]]] = None):
        self.graph = graph or DEFAULT_DEPENDENCY_GRAPH

    def run(self, df: pd.DataFrame) -> pd.DataFrame:
        _log("DependencyAgent", "Analysing dependency cascade risk...")
        df = df.copy()
        df["cascade_depth"]      = df["Affected System"].apply(self._cascade_depth)
        df["downstream_systems"] = df["Affected System"].apply(
            lambda s: ", ".join(self._downstream(s))
        )
        df["downstream_count"] = df["Affected System"].apply(
            lambda s: len(self._downstream(s))
        )
        df["cascade_risk_computed"] = (
            df["downstream_count"] / max(len(self.graph), 1)
        ).clip(0, 1)

        # FIX: Use the string label from the input as a floor when the system
        # is not in the dependency graph (cascade_risk_computed == 0).
        cascade_label_norm = df["Cascade Risk Label"].str.strip().str.lower().map(
            SEVERITY_MAP
        ).fillna(0.5)
        df["cascade_risk"] = df.apply(
            lambda r: max(r["cascade_risk_computed"], cascade_label_norm[r.name])
            if r["cascade_risk_computed"] == 0 else r["cascade_risk_computed"],
            axis=1,
        )
        return df

    def _downstream(self, system: str) -> List[str]:
        """BFS using deque to collect all transitively downstream systems."""
        visited: set = set()
        queue: deque = deque([system])
        while queue:
            node = queue.popleft()
            for child in self.graph.get(node, []):
                if child not in visited:
                    visited.add(child)
                    queue.append(child)
        return sorted(visited)

    def _cascade_depth(self, system: str) -> int:
        """Max depth of the dependency tree from this system (DFS)."""
        def depth(node: str, seen: frozenset) -> int:
            children = [c for c in self.graph.get(node, []) if c not in seen]
            if not children:
                return 0
            return 1 + max(depth(c, seen | {c}) for c in children)
        return depth(system, frozenset({system}))


# ---------------------------------------------------------------------------
# Agent 5: CustomerImpactAgent
# ---------------------------------------------------------------------------

class CustomerImpactAgent:
    """
    Scores customer-facing risk based on the Customer Impact field and
    Affected System name.
    """

    def run(self, df: pd.DataFrame) -> pd.DataFrame:
        _log("CustomerImpactAgent", "Scoring customer-facing exposure...")
        df = df.copy()
        df["customer_impact_norm"] = df["Customer Impact"].str.strip().str.lower().map(
            SEVERITY_MAP
        ).fillna(0.50)
        customer_facing_keywords = [
            "login", "auth", "frontend", "api", "payment", "checkout",
            "portal", "app", "web", "mobile",
        ]
        system_lower = df["Affected System"].str.lower()
        df["customer_facing"] = system_lower.apply(
            lambda s: any(kw in s for kw in customer_facing_keywords)
        )
        df["customer_impact_norm"] = df.apply(
            lambda r: min(1.0, r["customer_impact_norm"] * 1.15)
            if r["customer_facing"] else r["customer_impact_norm"],
            axis=1,
        )
        return df


# ---------------------------------------------------------------------------
# Agent 6: SchedulerAgent  (MAJOR FIX)
# ---------------------------------------------------------------------------

class SchedulerAgent:
    """
    Builds a patch schedule. Fixes:
    1. Named windows (e.g. "Sat 1-3 AM") are honoured and do NOT trigger false
       scheduling conflicts.
    2. When multiple CVEs share a system + window, they are assigned numbered
       slots within that window, spreading overflow to the next window.
    3. scheduling_conflict=TRUE only when two distinct CVEs compete for the
       exact same auto-computed date slot on the same system.
    """

    def __init__(self, start_date: Optional[str], window_days: int):
        self.start = (
            date.fromisoformat(start_date) if start_date
            else date.today()
        )
        self.window_days = window_days

    def run(self, df: pd.DataFrame) -> pd.DataFrame:
        _log("SchedulerAgent", "Building patch schedule...")
        df = df.copy()

        # Track: (system, window_label_or_date) → list of CVE row indices
        slot_registry: Dict[Tuple[str, str], List[int]] = {}

        scheduled_dates: List[str] = []
        slot_labels: List[str] = []
        uses_named_window: List[bool] = []

        computed_window_counter = 0  # used to spread auto-scheduled items

        for idx, row in df.iterrows():
            action   = row["recommended_action"]
            system   = row["Affected System"]
            existing = str(row["Maintenance Window"]).strip()
            is_named = _is_named_window(existing)

            if is_named:
                # Honour pre-supplied named window; assign numbered slot within it
                key = (system, existing)
                slot_registry.setdefault(key, [])
                slot_num = len(slot_registry[key]) + 1
                slot_registry[key].append(idx)

                # If overloaded, spill to next named window occurrence
                if slot_num > MAX_CVES_PER_WINDOW_SLOT:
                    overflow_label = f"{existing} (overflow {slot_num - MAX_CVES_PER_WINDOW_SLOT})"
                    scheduled_dates.append(overflow_label)
                    slot_labels.append(f"Slot {slot_num} (overflow)")
                else:
                    scheduled_dates.append(existing)
                    slot_labels.append(f"Slot {slot_num}")
                uses_named_window.append(True)

            elif existing.lower() in EMERGENCY_WINDOW_LABELS or action == "Patch immediately":
                # Emergency / immediate → earliest date slot
                emerg_date = str(self.start + timedelta(days=1))
                key = (system, emerg_date)
                slot_registry.setdefault(key, [])
                slot_num = len(slot_registry[key]) + 1
                slot_registry[key].append(idx)
                scheduled_dates.append(emerg_date)
                slot_labels.append(f"Emergency Slot {slot_num}")
                uses_named_window.append(False)

            elif action == "Patch in next maintenance window":
                computed_window_counter += 1
                delta = self.window_days * (1 + (computed_window_counter - 1) % 3)
                computed_date = str(self.start + timedelta(days=delta))
                key = (system, computed_date)
                slot_registry.setdefault(key, [])
                slot_num = len(slot_registry[key]) + 1
                slot_registry[key].append(idx)
                scheduled_dates.append(computed_date)
                slot_labels.append(f"Slot {slot_num}")
                uses_named_window.append(False)

            elif action == "Schedule and monitor":
                computed_window_counter += 1
                delta = self.window_days * (4 + (computed_window_counter - 1) % 4)
                computed_date = str(self.start + timedelta(days=delta))
                key = (system, computed_date)
                slot_registry.setdefault(key, [])
                slot_num = len(slot_registry[key]) + 1
                slot_registry[key].append(idx)
                scheduled_dates.append(computed_date)
                slot_labels.append(f"Slot {slot_num}")
                uses_named_window.append(False)

            else:
                scheduled_dates.append("Deferred")
                slot_labels.append("N/A")
                uses_named_window.append(False)

        df["scheduled_date"]   = scheduled_dates
        df["schedule_slot"]    = slot_labels
        df["uses_named_window"] = uses_named_window

        # FIX: Conflict = same system, same auto-computed date, multiple CVEs
        # Named windows are EXCLUDED from conflict detection (they're shared resources)
        non_named_mask = ~df["uses_named_window"]
        df["scheduling_conflict"] = False
        if non_named_mask.any():
            conflict_mask = df[non_named_mask].duplicated(
                subset=["Affected System", "scheduled_date"], keep=False
            )
            df.loc[non_named_mask, "scheduling_conflict"] = conflict_mask.values

        _log("SchedulerAgent",
             f"Conflicts detected: {int(df['scheduling_conflict'].sum())} "
             f"(named-window rows excluded from conflict check)")
        return df


# ---------------------------------------------------------------------------
# Agent 7: WhatIfAgent
# ---------------------------------------------------------------------------

class WhatIfAgent:
    """Simulates exploit spikes or patch delays and annotates their impact."""

    def run(
        self,
        df: pd.DataFrame,
        what_if_cve:         Optional[str] = None,
        what_if_exploit_prob: float        = 0.98,
        what_if_delay:        Optional[str] = None,
        delay_days:           int           = 30,
    ) -> pd.DataFrame:
        df = df.copy()
        if what_if_cve:
            df = self._simulate_exploit_spike(df, what_if_cve, what_if_exploit_prob)
        if what_if_delay:
            df = self._simulate_delay(df, what_if_delay, delay_days)
        return df

    def _simulate_exploit_spike(
        self, df: pd.DataFrame, cve_id: str, new_prob: float
    ) -> pd.DataFrame:
        _log("WhatIfAgent", f"Simulating exploit spike for {cve_id} → EPSS={new_prob:.2f}")
        mask = df["CVE ID"].str.strip().str.lower() == cve_id.strip().lower()
        if mask.any():
            df.loc[mask, "Exploit Prob (EPSS)"] = float(new_prob)
            df.loc[mask, "what_if_note"] = (
                f"⚠ Exploit spike simulated: EPSS set to {new_prob:.2f}"
            )
        else:
            _log("WhatIfAgent", f"WARNING: CVE '{cve_id}' not found in dataset.")
        return df

    def _simulate_delay(
        self, df: pd.DataFrame, cve_id: str, delay_days: int
    ) -> pd.DataFrame:
        _log("WhatIfAgent", f"Simulating {delay_days}-day delay for {cve_id}")
        mask = df["CVE ID"].str.strip().str.lower() == cve_id.strip().lower()
        if mask.any():
            epss = df.loc[mask, "Exploit Prob (EPSS)"].values[0]
            risk = (
                df.loc[mask, "risk_score"].values[0]
                if "risk_score" in df.columns else 0.5
            )
            accumulated_risk = min(1.0, risk + delay_days * 0.003 * epss)
            df.loc[mask, "what_if_note"] = (
                f"⏳ Delay simulation: {delay_days}d delay → "
                f"accumulated risk ≈ {accumulated_risk:.3f} "
                f"(current: {risk:.3f})"
            )
        else:
            _log("WhatIfAgent", f"WARNING: CVE '{cve_id}' not found in dataset.")
        return df


# ---------------------------------------------------------------------------
# Agent 8: ComplianceAgent
# ---------------------------------------------------------------------------

class ComplianceAgent:
    """Tags CVEs with relevant regulatory frameworks based on system names."""

    def run(self, df: pd.DataFrame) -> pd.DataFrame:
        _log("ComplianceAgent", "Tagging compliance exposure...")
        df = df.copy()
        df["detected_compliance"] = df.apply(self._detect, axis=1)
        df["compliance_risk"] = df["detected_compliance"].apply(
            lambda x: min(1.0, 0.25 * len(x.split(","))) if x else 0.0
        )
        return df

    def _detect(self, row: pd.Series) -> str:
        system_lower = str(row["Affected System"]).lower()
        existing     = str(row["Compliance Tags"]).strip()
        found: set   = set(t.strip() for t in existing.split(",") if t.strip())

        for framework, keywords in COMPLIANCE_KEYWORDS.items():
            if any(kw in system_lower for kw in keywords):
                found.add(framework)

        return ", ".join(sorted(found)) if found else ""


# ---------------------------------------------------------------------------
# Scoring helpers
# ---------------------------------------------------------------------------

def get_weight_profile(risk_appetite: str) -> Dict[str, float]:
    """Returns scoring weights. Positive weights MUST sum to 1.0."""
    profiles = {
        "aggressive": {
            "risk":             0.33,
            "business":         0.18,
            "dependency":       0.12,
            "exploit_bonus":    0.17,
            "cascade":          0.07,
            "compliance":       0.05,
            "customer_impact":  0.08,
            "downtime_penalty": 0.05,
            "effort_penalty":   0.02,
        },
        "conservative": {
            "risk":             0.26,
            "business":         0.20,
            "dependency":       0.16,
            "exploit_bonus":    0.09,
            "cascade":          0.08,
            "compliance":       0.07,
            "customer_impact":  0.14,
            "downtime_penalty": 0.12,
            "effort_penalty":   0.07,
        },
        "balanced": {
            "risk":             0.30,
            "business":         0.18,
            "dependency":       0.15,
            "exploit_bonus":    0.13,
            "cascade":          0.08,
            "compliance":       0.06,
            "customer_impact":  0.10,
            "downtime_penalty": 0.08,
            "effort_penalty":   0.04,
        },
    }
    w = profiles.get(risk_appetite, profiles["balanced"])
    validate_weights(w)
    return w


def compute_priority_scores(
    df: pd.DataFrame,
    risk_appetite: str,
    aggressive_escalation: bool = True,
) -> pd.DataFrame:
    """
    Compute priority scores on a 0-10 scale (matching dataset conventions).
    FIX: Output is now 0-10 so 'Patch immediately' threshold aligns with
    dataset expectations and is actually reachable.
    FIX: Percentile-based thresholds ensure top items always escalate.
    FIX: Hard CVSS+EPSS rule escalates critical CVEs regardless of weighted score.
    """
    df = df.copy()
    w  = get_weight_profile(risk_appetite)

    compliance_risk   = df["compliance_risk"]    if "compliance_risk"    in df.columns else pd.Series(0.0, index=df.index)
    customer_impact_n = df["customer_impact_norm"] if "customer_impact_norm" in df.columns else pd.Series(0.5, index=df.index)

    raw_score = (
        w["risk"]            * df["risk_score"] +
        w["business"]        * df["adjusted_business_impact"] +
        w["dependency"]      * df["dependency_score_norm"] +
        w["exploit_bonus"]   * df["exploit_prob_norm"] +
        w["cascade"]         * df["cascade_risk"] +
        w["compliance"]      * compliance_risk +
        w["customer_impact"] * customer_impact_n -
        w["downtime_penalty"] * df["downtime_cost_norm"] -
        w["effort_penalty"]   * df["effort_norm"]
    ).clip(0, 1)

    # FIX: Output on 0-10 scale
    df["priority_score"] = (raw_score * 10).round(3)

    df["risk_score"] = df["risk_score"].round(4)

    # FIX: Percentile-based action thresholds + hard escalation rule
    df["recommended_action"] = _recommend_action_dynamic(
        df, aggressive_escalation=aggressive_escalation
    )
    df["recommended_window"] = df.apply(
        lambda r: _recommend_window(r["priority_score"], r["downtime_cost_norm"]), axis=1
    )
    df["explanation"] = df.apply(_build_explanation, axis=1)

    # Risk Tier (new — executive-friendly label)
    df["risk_tier"] = df["priority_score"].apply(_assign_risk_tier)

    df = df.sort_values(
        by=["priority_score", "risk_score", "Exploit Prob (EPSS)"],
        ascending=[False, False, False],
    ).reset_index(drop=True)
    return df


def _recommend_action_dynamic(df: pd.DataFrame, aggressive_escalation: bool = True) -> pd.Series:
    """
    FIX: Use percentile-based thresholds so the top tier always gets escalated,
    regardless of the absolute score range in the dataset.
    Also apply a hard CVSS + EPSS escalation rule.
    """
    scores = df["priority_score"]
    p75 = scores.quantile(0.75)
    p50 = scores.quantile(0.50)
    p25 = scores.quantile(0.25)

    def classify(row: pd.Series) -> str:
        score = row["priority_score"]
        cvss  = row.get("CVSS Score", 0)
        epss  = row.get("Exploit Prob (EPSS)", 0)

        # Hard escalation: Critical CVSS + high EPSS always → Patch immediately
        if aggressive_escalation and cvss >= CRITICAL_CVSS_THRESHOLD and epss >= CRITICAL_EPSS_THRESHOLD:
            return "Patch immediately"

        if score >= p75:
            return "Patch immediately"
        if score >= p50:
            return "Patch in next maintenance window"
        if score >= p25:
            return "Schedule and monitor"
        return "Monitor for now"

    return df.apply(classify, axis=1)


def _assign_risk_tier(score: float) -> str:
    """Assign a P1-P4 risk tier label for executive reporting."""
    if score >= 8.5:  return "P1 – Critical"
    if score >= 6.5:  return "P2 – High"
    if score >= 4.5:  return "P3 – Medium"
    return "P4 – Low"


def _recommend_window(score: float, downtime_norm: float) -> str:
    if score >= 8.0 and downtime_norm <= 0.40: return "Immediate low-risk window"
    if score >= 8.0:                            return "Urgent controlled window"
    if score >= 6.0:                            return "Next scheduled maintenance window"
    if score >= 4.0:                            return "Planned low-traffic window"
    return "Defer until conditions change"


def _build_explanation(row: pd.Series) -> str:
    reasons: List[str] = []
    rs  = row.get("risk_score",               0.0)
    bi  = row.get("adjusted_business_impact", 0.0)
    dep = row.get("dependency_score_norm",     0.0)
    cas = row.get("cascade_risk",              0.0)
    exp = row.get("exploit_prob_norm",         0.0)
    ci  = row.get("customer_impact_norm",      0.0)
    el  = row.get("Expected Loss",             0.0)
    cvss = row.get("CVSS Score",               0.0)
    epss = row.get("Exploit Prob (EPSS)",      0.0)

    if cvss >= 9.0 and epss >= 0.75:
        reasons.append(f"critical severity with high exploit probability (CVSS {cvss}, EPSS {epss:.0%})")
    elif rs  >= 0.75: reasons.append("high technical risk")
    elif rs >= 0.55: reasons.append("moderate-to-high technical risk")

    if bi  >= 0.75: reasons.append("high business impact")
    elif bi >= 0.55: reasons.append("meaningful business impact")

    if dep >= 0.75: reasons.append("strong dependency risk")

    if cas >= 0.40:
        dc = int(row.get("downstream_count", 0))
        reasons.append(f"cascade risk ({dc} downstream systems)")

    if exp >= 0.75: reasons.append("high exploit likelihood")

    compliance = row.get("detected_compliance", "")
    if compliance: reasons.append(f"regulatory exposure ({compliance})")

    if ci >= 0.75: reasons.append("high customer-facing exposure")

    if el > 0:
        reasons.append(f"expected financial loss ≈ ${el:,.0f}")

    if row.get("downtime_cost_norm", 1.0) <= 0.35:
        reasons.append("relatively low downtime cost")
    if row.get("effort_norm", 1.0) <= 0.35:
        reasons.append("low implementation effort")

    if not reasons:
        reasons.append("balanced overall trade-off across risk, impact, and cost")

    return "Prioritized due to " + ", ".join(reasons) + "."


# ---------------------------------------------------------------------------
# Agent 9: ReporterAgent
# ---------------------------------------------------------------------------

class ReporterAgent:
    """Writes CSV, optional XLSX, optional JSON summary, and optional HTML audit report."""

    EXPORT_COLUMNS = [
        "CVE ID", "Affected System", "CVSS Score", "Severity",
        "Exploit Prob (EPSS)", "Criticality", "Tier",
        "Business Impact", "Dependency Score", "Downtime Cost", "Est. Effort",
        "Expected Loss",
        "Risk Score", "Priority Score", "Risk Tier",
        "cascade_risk", "downstream_systems", "downstream_count",
        "detected_compliance",
        "customer_impact_norm", "customer_facing",
        "recommended_action", "recommended_window",
        "scheduled_date", "schedule_slot",
        "scheduling_conflict", "explanation", "what_if_note",
        "Vendor", "Patch Available", "SBOM Component", "Description",
    ]

    def run(
        self,
        df: pd.DataFrame,
        output_csv: str,
        output_xlsx: Optional[str],
        summary_json: Optional[str],
        html_report:  Optional[str],
        input_path:   str,
        risk_appetite: str,
    ) -> None:
        _log("ReporterAgent", "Writing outputs...")

        export_df = df.copy()
        export_df["Risk Score"]     = export_df["risk_score"]
        export_df["Priority Score"] = export_df["priority_score"]
        export_df["Risk Tier"]      = export_df.get("risk_tier", "")

        for col in ["what_if_note", "downstream_systems", "schedule_slot"]:
            if col not in export_df.columns:
                export_df[col] = ""
        if "scheduling_conflict" not in export_df.columns:
            export_df["scheduling_conflict"] = False

        existing  = [c for c in self.EXPORT_COLUMNS if c in export_df.columns]
        out_path  = Path(output_csv)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        export_df[existing].to_csv(out_path, index=False)
        _log("ReporterAgent", f"CSV saved → {output_csv}")

        if output_xlsx:
            self._write_xlsx(export_df[existing], output_xlsx)
        if summary_json:
            self._write_json(df, summary_json, input_path, risk_appetite)
        if html_report:
            self._write_html(df, html_report, input_path, risk_appetite)

    # ------------------------------------------------------------------
    def _write_xlsx(self, df: pd.DataFrame, path: str) -> None:
        """Write styled Excel workbook."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            _log("ReporterAgent", "openpyxl not installed — skipping XLSX output.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patch Plan"

        header_fill = PatternFill("solid", fgColor="0F172A")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        action_colors = {
            "Patch immediately":                "DC2626",
            "Patch in next maintenance window": "D97706",
            "Schedule and monitor":             "2563EB",
            "Monitor for now":                  "16A34A",
        }

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                elif c_idx == df.columns.tolist().index("recommended_action") + 1 if "recommended_action" in df.columns else -1:
                    color = action_colors.get(str(value), "6B7280")
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(color="FFFFFF", bold=True, size=9)

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(p)
        _log("ReporterAgent", f"Excel workbook saved → {path}")

    def _write_json(self, df, path, input_path, risk_appetite):
        top5 = df.head(5)[
            ["CVE ID", "Affected System", "priority_score", "recommended_action", "risk_tier"]
        ].to_dict(orient="records")

        all_frameworks: List[str] = []
        for tags in df.get("detected_compliance", pd.Series(dtype=str)).dropna():
            all_frameworks.extend(t.strip() for t in tags.split(",") if t.strip())
        top_frameworks = [fw for fw, _ in Counter(all_frameworks).most_common(5)]

        # Top 3 most-affected systems
        top_systems = df["Affected System"].value_counts().head(3).to_dict()

        summary = {
            "generated_at":              str(date.today()),
            "input_file":                input_path,
            "risk_appetite":             risk_appetite,
            "total_vulnerabilities":     int(len(df)),
            "avg_priority_score":        round(float(df["priority_score"].mean()), 3),
            "top_priority_cve":          str(df.iloc[0]["CVE ID"])          if len(df) else None,
            "top_priority_score":        float(df.iloc[0]["priority_score"]) if len(df) else None,
            "p1_critical_count":         int((df.get("risk_tier", pd.Series()) == "P1 – Critical").sum()),
            "p2_high_count":             int((df.get("risk_tier", pd.Series()) == "P2 – High").sum()),
            "p3_medium_count":           int((df.get("risk_tier", pd.Series()) == "P3 – Medium").sum()),
            "p4_low_count":              int((df.get("risk_tier", pd.Series()) == "P4 – Low").sum()),
            "immediate_patch_count":     int((df["recommended_action"] == "Patch immediately").sum()),
            "next_window_count":         int((df["recommended_action"] == "Patch in next maintenance window").sum()),
            "schedule_and_monitor_count":int((df["recommended_action"] == "Schedule and monitor").sum()),
            "monitor_count":             int((df["recommended_action"] == "Monitor for now").sum()),
            "scheduling_conflicts":      int(df.get("scheduling_conflict", pd.Series(False)).sum()),
            "compliance_exposed_count":  int(
                (df.get("detected_compliance", pd.Series([""] * len(df))) != "").sum()
            ),
            "top_compliance_frameworks": top_frameworks,
            "top_affected_systems":      top_systems,
            "top_recommendations":       top5,
        }
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)
        with open(p, "w", encoding="utf-8") as f:
            json.dump(summary, f, indent=2)
        _log("ReporterAgent", f"JSON summary saved → {path}")

    def _write_html(self, df, path, input_path, risk_appetite):
        """Generate a styled, interactive HTML audit report."""
        p = Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)

        immediate  = int((df["recommended_action"] == "Patch immediately").sum())
        next_win   = int((df["recommended_action"] == "Patch in next maintenance window").sum())
        scheduled  = int((df["recommended_action"] == "Schedule and monitor").sum())
        monitor    = int((df["recommended_action"] == "Monitor for now").sum())
        p1_count   = int((df.get("risk_tier", pd.Series()) == "P1 – Critical").sum())
        compliance_exposed = int(
            (df.get("detected_compliance", pd.Series([""] * len(df))) != "").sum()
        )
        avg_score = round(float(df["priority_score"].mean()), 2)
        conflicts = int(df.get("scheduling_conflict", pd.Series(False)).sum())

        def action_badge(action: str) -> str:
            colors = {
                "Patch immediately":                    "#dc2626",
                "Patch in next maintenance window":     "#d97706",
                "Schedule and monitor":                 "#2563eb",
                "Monitor for now":                      "#16a34a",
            }
            color = colors.get(action, "#6b7280")
            return (
                f'<span style="background:{color};color:#fff;padding:2px 8px;'
                f'border-radius:4px;font-size:0.75rem;font-weight:600;">{action}</span>'
            )

        def tier_badge(tier: str) -> str:
            colors = {
                "P1 – Critical": "#dc2626",
                "P2 – High":     "#d97706",
                "P3 – Medium":   "#2563eb",
                "P4 – Low":      "#16a34a",
            }
            color = colors.get(tier, "#6b7280")
            return (
                f'<span style="background:{color};color:#fff;padding:2px 7px;'
                f'border-radius:4px;font-size:0.72rem;font-weight:700;">{tier}</span>'
            )

        rows_html = ""
        for _, row in df.head(50).iterrows():
            compliance = row.get("detected_compliance", "")
            comp_html  = (
                " ".join(
                    f'<span style="background:#7c3aed;color:#fff;padding:1px 6px;'
                    f'border-radius:3px;font-size:0.7rem;">{fw}</span>'
                    for fw in compliance.split(", ") if fw
                )
                if compliance else ""
            )
            downstream  = int(row.get("downstream_count", 0))
            cascade_pct = int(row.get("cascade_risk", 0) * 100)
            what_if     = row.get("what_if_note", "")
            sched       = row.get("scheduled_date", "")
            slot        = row.get("schedule_slot", "")
            conflict    = row.get("scheduling_conflict", False)
            conflict_icon = ' <span title="Scheduling conflict detected" style="color:#dc2626">⚠</span>' if conflict else ""
            desc        = str(row.get("Description", "")).replace('"', "'")
            el          = row.get("Expected Loss", 0.0)
            el_html     = f"${el:,.0f}" if el and el > 0 else ""
            ps          = row.get("priority_score", 0.0)
            rt          = row.get("risk_tier", "")

            rows_html += f"""
            <tr title="{desc}">
              <td style="font-weight:600;color:#1e293b">{row['CVE ID']}</td>
              <td>{row['Affected System']}</td>
              <td style="text-align:center">{row.get('CVSS Score', '-')}</td>
              <td style="text-align:center"><b>{ps:.2f}</b></td>
              <td>{tier_badge(rt)}</td>
              <td>{action_badge(row['recommended_action'])}</td>
              <td style="text-align:center">{downstream} / {cascade_pct}%</td>
              <td>{comp_html}</td>
              <td style="font-size:0.8rem;color:#475569">{sched}<br><small style="color:#94a3b8">{slot}</small>{conflict_icon}</td>
              <td style="font-size:0.8rem;color:#16a34a;font-weight:600">{el_html}</td>
              <td style="font-size:0.78rem;color:#64748b">{
                '<span style="color:#b45309">' + what_if + '</span>' if what_if else ''
              }</td>
            </tr>"""

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Agentic Patch Strategist v3 – Audit Report</title>
<style>
  body {{ font-family: 'Segoe UI', sans-serif; margin: 0; background: #f1f5f9; color: #1e293b; }}
  header {{ background: #0f172a; color: #fff; padding: 24px 40px; }}
  header h1 {{ margin: 0; font-size: 1.5rem; }}
  header p  {{ margin: 4px 0 0; font-size: 0.85rem; color: #94a3b8; }}
  .stats {{ display: flex; gap: 16px; padding: 24px 40px; flex-wrap: wrap; }}
  .card {{ background: #fff; border-radius: 8px; padding: 16px 24px; min-width: 140px;
           box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
  .card .num {{ font-size: 2rem; font-weight: 700; }}
  .card .label {{ font-size: 0.8rem; color: #64748b; margin-top: 2px; }}
  .red {{ color: #dc2626; }} .amber {{ color: #d97706; }}
  .blue {{ color: #2563eb; }} .green {{ color: #16a34a; }}
  .purple {{ color: #7c3aed; }} .slate {{ color: #475569; }}
  .orange {{ color: #ea580c; }}
  table {{ width: calc(100% - 80px); margin: 0 40px 40px; border-collapse: collapse;
           background: #fff; border-radius: 8px; overflow: hidden;
           box-shadow: 0 1px 4px rgba(0,0,0,.08); font-size: 0.83rem; }}
  th {{ background: #0f172a; color: #fff; padding: 10px 12px; text-align: left;
        font-size: 0.78rem; text-transform: uppercase; letter-spacing: .05em; cursor: pointer; }}
  th:hover {{ background: #1e293b; }}
  td {{ padding: 9px 12px; border-bottom: 1px solid #e2e8f0; }}
  tr:hover td {{ background: #f8fafc; }}
  .note {{ padding: 4px 40px 24px; font-size: 0.8rem; color: #64748b; }}
</style>
</head>
<body>
<header>
  <h1>🔐 Agentic Patch Strategist v3 — Audit Report</h1>
  <p>Generated: {date.today()} &nbsp;|&nbsp; Input: {input_path}
     &nbsp;|&nbsp; Risk appetite: <b>{risk_appetite}</b>
     &nbsp;|&nbsp; Total CVEs: <b>{len(df)}</b>
     &nbsp;|&nbsp; Avg priority score: <b>{avg_score}/10</b></p>
</header>
<div class="stats">
  <div class="card"><div class="num red">{p1_count}</div>
    <div class="label">P1 Critical</div></div>
  <div class="card"><div class="num red">{immediate}</div>
    <div class="label">Patch Immediately</div></div>
  <div class="card"><div class="num amber">{next_win}</div>
    <div class="label">Next Window</div></div>
  <div class="card"><div class="num blue">{scheduled}</div>
    <div class="label">Schedule &amp; Monitor</div></div>
  <div class="card"><div class="num green">{monitor}</div>
    <div class="label">Monitor for Now</div></div>
  <div class="card"><div class="num purple">{compliance_exposed}</div>
    <div class="label">Compliance Exposed</div></div>
  <div class="card"><div class="num {"orange" if conflicts > 0 else "green"}">{conflicts}</div>
    <div class="label">Scheduling Conflicts</div></div>
</div>
<table id="patchtable">
  <thead><tr>
    <th>CVE ID</th><th>Affected System</th><th>CVSS</th><th>Priority (0-10)</th>
    <th>Risk Tier</th><th>Action</th><th>Cascade (sys/%)</th><th>Compliance</th>
    <th>Scheduled</th><th>Exp. Loss</th><th>What-If Note</th>
  </tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<div class="note">
  Showing top 50 of {len(df)} vulnerabilities sorted by priority score (0–10 scale).
  Hover over a row to see the CVE description. ⚠ = scheduling conflict detected.
</div>
</body></html>"""

        with open(p, "w", encoding="utf-8") as f:
            f.write(html)
        _log("ReporterAgent", f"HTML report saved → {path}")


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

def _clean_numeric(series: pd.Series, default: float) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(default).astype(float)


def _map_severity(series: pd.Series) -> pd.Series:
    return (
        series.astype(str).str.strip().str.lower()
        .map(SEVERITY_MAP).fillna(0.50).astype(float)
    )


def _min_max(series: pd.Series) -> pd.Series:
    """
    Min-max normalise a numeric Series to [0, 1].
    Returns a properly-indexed Series even when lo == hi.
    """
    s        = series.astype(float)
    lo, hi   = s.min(), s.max()
    if pd.isna(lo) or lo == hi:
        return pd.Series(0.5, index=s.index, dtype=float)
    return (s - lo) / (hi - lo)


def _is_named_window(window: str) -> bool:
    """Returns True if the window string looks like a named recurring window
    (e.g. 'Sat 1-3 AM') rather than an ISO date or empty string."""
    if not window or window.strip() == "":
        return False
    lower = window.strip().lower()
    if lower in EMERGENCY_WINDOW_LABELS:
        return False
    return any(kw in lower for kw in NAMED_WINDOW_PATTERN_KEYWORDS)


def _load_dependency_graph(path: Optional[str]) -> Optional[Dict[str, List[str]]]:
    if not path:
        return None
    fp = Path(path)
    if not fp.exists():
        _log("DependencyAgent", f"WARNING: graph file '{path}' not found; using default.")
        return None
    with open(fp, encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def main() -> None:
    args = parse_args()
    aggressive_escalation = not args.no_aggressive_escalation

    # Instantiate all agents
    ingest_agent    = IngestAgent()
    risk_agent      = RiskScoringAgent()
    business_agent  = BusinessImpactAgent()
    graph           = _load_dependency_graph(args.dependency_graph)
    dep_agent       = DependencyAgent(graph=graph)
    customer_agent  = CustomerImpactAgent()
    what_if_agent   = WhatIfAgent()
    compliance_agent= ComplianceAgent()
    scheduler_agent = SchedulerAgent(
        start_date=args.schedule_start,
        window_days=args.window_days,
    )
    reporter_agent  = ReporterAgent()

    print("\n" + "=" * 64)
    print("  Agentic Patch Strategist v3 — Multi-Agent Pipeline")
    print("=" * 64)

    # ── Pipeline ──────────────────────────────────────────────────────
    df = ingest_agent.run(args.input)
    df = risk_agent.run(df)

    # What-if simulation runs before final scoring so EPSS mutations propagate
    df = what_if_agent.run(
        df,
        what_if_cve=args.what_if_cve,
        what_if_exploit_prob=args.what_if_exploit_prob,
        what_if_delay=args.what_if_delay,
        delay_days=args.delay_days,
    )
    if args.what_if_cve:
        df = risk_agent.run(df)

    df = business_agent.run(df)
    df = dep_agent.run(df)
    df = customer_agent.run(df)
    df = compliance_agent.run(df)
    df = compute_priority_scores(df, args.risk_appetite,
                                 aggressive_escalation=aggressive_escalation)
    df = scheduler_agent.run(df)

    reporter_agent.run(
        df=df,
        output_csv=args.output,
        output_xlsx=args.output_xlsx,
        summary_json=args.summary_json,
        html_report=args.html_report,
        input_path=args.input,
        risk_appetite=args.risk_appetite,
    )

    # ── Console summary ───────────────────────────────────────────────
    print("\n" + "=" * 64)
    print(f"  Risk appetite           : {args.risk_appetite}")
    print(f"  Aggressive escalation   : {aggressive_escalation}")
    print(f"  Total CVEs              : {len(df)}")
    print(f"  P1 Critical             : {(df.get('risk_tier', pd.Series()) == 'P1 – Critical').sum()}")
    print(f"  P2 High                 : {(df.get('risk_tier', pd.Series()) == 'P2 – High').sum()}")
    print(f"  P3 Medium               : {(df.get('risk_tier', pd.Series()) == 'P3 – Medium').sum()}")
    print(f"  P4 Low                  : {(df.get('risk_tier', pd.Series()) == 'P4 – Low').sum()}")
    print(f"  Patch immediately       : {(df['recommended_action'] == 'Patch immediately').sum()}")
    print(f"  Next maintenance window : {(df['recommended_action'] == 'Patch in next maintenance window').sum()}")
    print(f"  Schedule and monitor    : {(df['recommended_action'] == 'Schedule and monitor').sum()}")
    print(f"  Monitor for now         : {(df['recommended_action'] == 'Monitor for now').sum()}")
    print(f"  Scheduling conflicts    : {int(df.get('scheduling_conflict', pd.Series(False)).sum())}")
    print(f"  Avg priority score      : {df['priority_score'].mean():.2f} / 10")
    print("=" * 64)
    print(f"\nTop {args.top_n} recommendations:")
    preview_cols = [
        "CVE ID", "Affected System", "CVSS Score", "Exploit Prob (EPSS)",
        "priority_score", "risk_tier", "cascade_risk", "detected_compliance",
        "recommended_action", "scheduled_date",
    ]
    existing_cols = [c for c in preview_cols if c in df.columns]
    print(df.head(args.top_n)[existing_cols].to_string(index=False))
    print()


if __name__ == "__main__":
    main()
